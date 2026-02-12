"""Excel output writer module - generates a 2-sheet Excel file."""

import logging
from decimal import Decimal
from pathlib import Path

from openpyxl import Workbook
from openpyxl.styles import Alignment, Border, Font, PatternFill, Side
from openpyxl.utils import get_column_letter

from .schema import ExtractedInvoice

logger = logging.getLogger(__name__)

# Formatting constants
HEADER_FONT = Font(bold=True, size=11, color="000000")
HEADER_FILL = PatternFill(start_color="D9E1F2", end_color="D9E1F2", fill_type="solid")
ALT_ROW_FILL = PatternFill(start_color="E8F0FE", end_color="E8F0FE", fill_type="solid")
HEADER_BORDER = Border(bottom=Side(style="thin", color="000000"))
RIGHT_ALIGN = Alignment(horizontal="right", vertical="center")
LEFT_ALIGN = Alignment(horizontal="left", vertical="center")

# Number formats
MONEY_FORMAT = '#,##0.00'
PERCENT_FORMAT = '0.00"%"'

# Column definitions for CGST/SGST invoices (Zomato)
CGST_SGST_COLUMNS = [
    {"header": "Description", "field": "description", "width": 35, "format": None, "align": "left"},
    {"header": "Gross", "field": "gross_value", "width": 14, "format": MONEY_FORMAT, "align": "right"},
    {"header": "Discount", "field": "discount", "width": 14, "format": MONEY_FORMAT, "align": "right"},
    {"header": "Net", "field": "net_value", "width": 14, "format": MONEY_FORMAT, "align": "right"},
    {"header": "CGST Rate", "field": "cgst_rate", "width": 12, "format": PERCENT_FORMAT, "align": "right"},
    {"header": "CGST Amount", "field": "cgst_amount", "width": 14, "format": MONEY_FORMAT, "align": "right"},
    {"header": "SGST Rate", "field": "sgst_rate", "width": 12, "format": PERCENT_FORMAT, "align": "right"},
    {"header": "SGST Amount", "field": "sgst_amount", "width": 14, "format": MONEY_FORMAT, "align": "right"},
    {"header": "Total", "field": "total", "width": 14, "format": MONEY_FORMAT, "align": "right"},
]

# Column definitions for IGST invoices (Flipkart)
IGST_COLUMNS = [
    {"header": "Description", "field": "description", "width": 40, "format": None, "align": "left"},
    {"header": "Gross", "field": "gross_value", "width": 14, "format": MONEY_FORMAT, "align": "right"},
    {"header": "Discount", "field": "discount", "width": 14, "format": MONEY_FORMAT, "align": "right"},
    {"header": "Taxable Value", "field": "net_value", "width": 14, "format": MONEY_FORMAT, "align": "right"},
    {"header": "IGST Rate", "field": "igst_rate", "width": 12, "format": PERCENT_FORMAT, "align": "right"},
    {"header": "IGST Amount", "field": "igst_amount", "width": 14, "format": MONEY_FORMAT, "align": "right"},
    {"header": "CESS", "field": "cess_amount", "width": 14, "format": MONEY_FORMAT, "align": "right"},
    {"header": "Total", "field": "total", "width": 14, "format": MONEY_FORMAT, "align": "right"},
]


def write_excel(invoice: ExtractedInvoice, output_path: str) -> str:
    """
    Write a structured 2-sheet Excel file from the extracted invoice data.

    Sheet 1: Invoice Summary (key-value layout)
    Sheet 2: Line Items (tabular layout)

    Args:
        invoice: ExtractedInvoice object with all extracted data.
        output_path: Path for the output Excel file.

    Returns:
        The output file path.
    """
    output_dir = Path(output_path).parent
    output_dir.mkdir(parents=True, exist_ok=True)

    wb = Workbook()

    # --- Sheet 1: Invoice Summary ---
    ws_summary = wb.active
    ws_summary.title = "Invoice Summary"
    ws_summary.sheet_properties.tabColor = "4472C4"

    _write_summary_sheet(ws_summary, invoice)

    # --- Sheet 2: Line Items ---
    ws_items = wb.create_sheet(title="Line Items")
    ws_items.sheet_properties.tabColor = "70AD47"

    # Choose column layout based on tax type
    columns = IGST_COLUMNS if invoice.tax_type == "igst" else CGST_SGST_COLUMNS
    _write_line_items_sheet(ws_items, invoice, columns)

    wb.save(output_path)
    logger.info(f"Excel file written to: {output_path}")

    return output_path


def _write_summary_sheet(ws, invoice: ExtractedInvoice) -> None:
    """Write the Invoice Summary sheet (key-value layout)."""
    ws.column_dimensions["A"].width = 25
    ws.column_dimensions["B"].width = 40

    ws["A1"] = "Invoice Summary"
    ws["A1"].font = Font(bold=True, size=14, color="4472C4")
    ws.merge_cells("A1:B1")

    summary_data = [
        ("Invoice Number", invoice.invoice_number),
        ("Invoice Date", invoice.invoice_date.strftime("%d/%m/%Y")),
        ("Vendor Name", invoice.vendor_name),
        ("Vendor GST", invoice.vendor_gst),
        ("Customer Name", invoice.customer_name),
        ("State", invoice.state),
        ("Grand Total (Raw)", float(invoice.grand_total_raw)),
        ("Grand Total (Rounded)", float(invoice.grand_total_rounded)),
    ]

    # Add optional Flipkart fields
    if invoice.order_id:
        summary_data.insert(1, ("Order ID", invoice.order_id))

    for i, (label, value) in enumerate(summary_data, start=3):
        ws.cell(row=i, column=1, value=label)
        ws.cell(row=i, column=1).font = Font(bold=True, size=11)
        ws.cell(row=i, column=1).fill = HEADER_FILL

        cell = ws.cell(row=i, column=2, value=value)
        if isinstance(value, float):
            cell.number_format = MONEY_FORMAT
            cell.alignment = RIGHT_ALIGN
        else:
            cell.alignment = LEFT_ALIGN


def _write_line_items_sheet(ws, invoice: ExtractedInvoice, columns: list) -> None:
    """Write the Line Items sheet (tabular layout)."""
    # Write headers
    for col_idx, col_config in enumerate(columns, start=1):
        cell = ws.cell(row=1, column=col_idx, value=col_config["header"])
        cell.font = HEADER_FONT
        cell.fill = HEADER_FILL
        cell.border = HEADER_BORDER
        col_letter = get_column_letter(col_idx)
        ws.column_dimensions[col_letter].width = col_config["width"]

        if col_config["align"] == "right":
            cell.alignment = RIGHT_ALIGN
        else:
            cell.alignment = LEFT_ALIGN

    # Write data rows
    for row_idx, item in enumerate(invoice.line_items, start=2):
        for col_idx, col_config in enumerate(columns, start=1):
            value = getattr(item, col_config["field"])

            if col_config["format"] is not None and col_config["field"] != "description":
                value = float(value)

            cell = ws.cell(row=row_idx, column=col_idx, value=value)

            if col_config["format"]:
                cell.number_format = col_config["format"]

            if col_config["align"] == "right":
                cell.alignment = RIGHT_ALIGN
            else:
                cell.alignment = LEFT_ALIGN

            if row_idx % 2 == 0:
                cell.fill = ALT_ROW_FILL

    # Auto-filter
    last_col_letter = get_column_letter(len(columns))
    last_row = len(invoice.line_items) + 1
    ws.auto_filter.ref = f"A1:{last_col_letter}{last_row}"
